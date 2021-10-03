from numpy.core.fromnumeric import size
from openpyxl import load_workbook
from openpyxl.styles import Font
from sklearn import tree
import Calculate as CA
from RandomCal import RandomGenerate
import data_process
import RandomCal
from itertools import permutations
import os
import numpy as np
import pandas as pd
import sys
from sklearn import preprocessing
from sklearn.preprocessing import LabelEncoder

from sklearn.tree import DecisionTreeClassifier
import statistics
from sklearn.metrics import roc_auc_score
from sklearn import svm
import time
import datetime

# 每一種方式不同比例中最大準確率的改成紅色 todo
# 計算時間 不同方法 不同比例的執行時間


'''give the path'''


def train_test_split(folder):
    train = []
    test = []
    # os.chdir(folder)
    dirs = os.listdir(folder)
    print('dirss',dirs,'folder',folder)
    for i in dirs:
        # print(i.split("-")[-1])
        if("xlsx" in i):
            if("tra" in i):
                train.append(i)

            elif("tst" in i):
                test.append(i)
    train = sorted(train)
    test = sorted(test)
    return train, test


'''2 menthods merge with different ration'''


def calculatetwomethod(train, test, row, path, approach, maxsingle, sheetName):
    '''
   1. train is assign by train data
   2. test
   3. approach means the choice of random or center
   4. id is for the purpose of writing into the excel different
       dataset need to be on different row
   5. path right now is useless
   '''
  # 思路 2個方法混 --> 3個加進去 其中一個比例為 0
    cell = 1
    originpath = os.getcwd()
    os.chdir(path)
    tempfileName = train[0]  # 要擷取 file 的名字
    tempfileName = tempfileName.split('-')
    fileName = '-'.join(tempfileName[:-2])
    data = pd.read_excel(train[0], index_col=0)
    # print("dsdsadad", data.columns)

    randompoly = []
    randomPro = []
    randomIPF = []
    start = 0
    end = 0
    tempappr = []  # 放要比較的數值(5,5,0 之間會比較)
    for index in range(len(m)):

        if approach == 'random':
            print('m[index]', m[index])
            randompoly = RandomCal.RandomGenerate(
                train, int(m[index][0])*0.1, "poly", path)
            randomPro = RandomCal.RandomGenerate(
                train, int(m[index][1])*0.1, "prow", path)
            randomIPF = RandomCal.RandomGenerate(
                train, int(m[index][2])*0.1, "SMOTEIPF", path)
        elif approach == 'center':
            randompoly = RandomCal.CenterGenerate(
                train, int(m[index][0])*0.1, "poly", path)
            randomPro = RandomCal.CenterGenerate(
                train, int(m[index][1])*0.1, "prow", path)
            randomIPF = RandomCal.CenterGenerate(
                train, int(m[index][2])*0.1, "SMOTEIPF", path)
        elif approach == 'elbowRandom':
            randompoly = RandomCal.ElbowRandomGenerate(
                train, int(m[index][0])*0.1, "poly", path)
            randomPro = RandomCal.ElbowRandomGenerate(
                train, int(m[index][1])*0.1, "prow", path)
            randomIPF = RandomCal.ElbowRandomGenerate(
                train, int(m[index][2])*0.1, "SMOTEIPF", path)
        elif approach == 'elbowCenter':
            randompoly = RandomCal.ElbowCenterGenerate(
                train, int(m[index][0])*0.1, "poly", path)
            randomPro = RandomCal.ElbowCenterGenerate(
                train, int(m[index][1])*0.1, "prow", path)
            randomIPF = RandomCal.ElbowCenterGenerate(
                train, int(m[index][2])*0.1, "SMOTEIPF", path)
        # start = datetime.datetime.now()
        start = time.process_time()

        allRandom = []
        temp = []
        maxmethod = 0
        maxmethod = max(len(randomPro), len(randomIPF), len(randompoly))
        for i in range(maxmethod):
            if len(randomPro) == 0:
                temp = randomIPF[i] + randompoly[i]
            elif len(randomIPF) == 0:
                temp = randompoly[i] + randomPro[i]
            elif len(randompoly) == 0:
                temp = randomIPF[i] + randomPro[i]
            # temp = randomIPF[i] + randomPro[i]  # list 合併
            # temp = temp + randompoly[i]
            temp = np.array(temp)
            allRandom.append(temp)

        for j in range(len(allRandom)):  #
            allRandom[j] = pd.DataFrame(allRandom[j], columns=data.columns)

        cell = cell + 1
        # 為了算個別 C4.5 跟 SVM 時間 所以先註解掉 SVM 的部分
        meanDe = CA.predictDe(train, test, allRandom)
        # meanSVM = CA.predictSVM(train, test, allRandom)

        end = time.process_time()

        # write in excel prepare
        # 時間
        dur = end - start

        wb = load_workbook(r'C:\Users\lab723\Desktop\Research\test2-1.xlsx')
        # sheet = wb['Ensemble 三個方法 Random']
        # sheet Name
        sheet = wb[sheetName]
        sheet.cell(row=1+id, column=1, value=fileName)
        sheet.cell(row=1+id, column=cell, value=meanDe)
        # sheet.cell(row=3+id, column=cell, value=meanSVM)
        # 時間
        sheet.cell(row=2+id, column=cell, value=dur)

        # 找最大值 並上紅色
        ma = meanDe
        # ma = meanSVM
        tempappr.append(ma)

        fontRed = Font(color='FF0000')  # point at red font
        fontBoldRed = Font(color='FF0000', bold=True)
        if ma > maxsingle:

            sheet.cell(1+id, cell).font = fontBoldRed  # C4.5
            # sheet.cell(3+id, cell).font = fontBoldRed  # SVM

            print('meanDe233', ma)
        if index == 2 or index == 8 or index == 14:
            maxcell = max(tempappr)  # 每個方法的最大值
            if maxcell < maxsingle:
                print("new round3", tempappr)
                maxindex = tempappr.index(maxcell)  # 最大值的 index
            # print('maxindex', maxindex)
                # r = 3+id  # SVM
                r = 1 + id  # C4.5
                c = cell - (len(tempappr) - maxindex) + 1
                print('row3', r, '  column4', c)
                sheet.cell(r, c).font = fontRed

                print('meanDe111', maxcell)
            tempappr = []
        # 若大於 singel method 則標粗體

        wb.save(r'C:\Users\lab723\Desktop\Research\test2-1.xlsx')
    os.chdir(originpath)


'''3 method merge with different ratio'''


def calculatethreemethod(train, test, id, path, approach, maxsingle, sheetName):
    '''
    1. train is assign by train data
    2. test
    3. approach means the choice of random or center
    4. id is for the purpose of writing into the excel different
        dataset need to be on different row
    5. path right now is useless
    '''
    cell = 1
    originpath = os.getcwd()
    
    os.chdir(path)
    tempfileName = train[0]  # 要擷取 file 的名字
    tempfileName = tempfileName.split('-')
    fileName = '-'.join(tempfileName[:-2])
    data = pd.read_excel(train[0], index_col=0)
    # print("dsdsadad", data.columns)
    tempappr = []  # 放要比較的數值(3,3,4 之間會比較)
    randompoly = []
    randomPro = []
    randomIPF = []
    start = 0
    end = 0
    for index in range(len(f)):

        if approach == 'random':
            randompoly = RandomCal.RandomGenerate(
                train, int(f[index][0])*0.1, "poly", path)
            randomPro = RandomCal.RandomGenerate(
                train, int(f[index][1])*0.1, "prow", path)
            randomIPF = RandomCal.RandomGenerate(
                train, int(f[index][2])*0.1, "SMOTEIPF", path)
        elif approach == 'center':
            randompoly = RandomCal.CenterGenerate(
                train, int(f[index][0])*0.1, "poly", path)
            randomPro = RandomCal.CenterGenerate(
                train, int(f[index][1])*0.1, "prow", path)
            randomIPF = RandomCal.CenterGenerate(
                train, int(f[index][2])*0.1, "SMOTEIPF", path)
        elif approach == 'elbowRandom':
            randompoly = RandomCal.ElbowRandomGenerate(
                train, int(f[index][0])*0.1, "poly", path)
            randomPro = RandomCal.ElbowRandomGenerate(
                train, int(f[index][1])*0.1, "prow", path)
            randomIPF = RandomCal.ElbowRandomGenerate(
                train, int(f[index][2])*0.1, "SMOTEIPF", path)
        elif approach == 'elbowCenter':
            randompoly = RandomCal.ElbowCenterGenerate(
                train, int(f[index][0])*0.1, "poly", path)
            randomPro = RandomCal.ElbowCenterGenerate(
                train, int(f[index][1])*0.1, "prow", path)
            randomIPF = RandomCal.ElbowCenterGenerate(
                train, int(f[index][2])*0.1, "SMOTEIPF", path)
        # start = datetime.datetime.now()
        start = time.process_time()

        allRandom = []
        temp = []
        for i in range(len(randomPro)):
            temp = randomIPF[i] + randomPro[i]  # list 合併
            temp = temp + randompoly[i]
            temp = np.array(temp)
            allRandom.append(temp)

        for j in range(len(allRandom)):  #
            allRandom[j] = pd.DataFrame(allRandom[j], columns=data.columns)

        cell = cell + 1
        # 為了算個別 C4.5 跟 SVM 時間 所以先註解掉 SVM 的部分
        # meanDe = CA.predictDe(train, test, allRandom)
        meanSVM = CA.predictSVM(train, test, allRandom)
        end = time.process_time()

        # write in excel prepare
        # 時間
        dur = end - start
        dur = round(dur, 3)
        wb = load_workbook(r'C:\Users\lab723\Desktop\Research\test2-1.xlsx')
        # sheet = wb['Ensemble 三個方法 Random']
        # sheet = wb['Ensemble 三個方法 ElbowCenter']
        # sheet Name
        sheet = wb[sheetName]
        sheet.cell(row=1+id, column=1, value=fileName)
        # sheet.cell(row=1+id, column=cell, value=meanDe)
        # 時間

        sheet.cell(row=4+id, column=cell, value=dur) # need to be modified to --> 4+id  when running the svm
        sheet.cell(row=3+id, column=cell, value=meanSVM)

        # 找最大值 並上紅色
        # ma = meanDe
        ma = meanSVM
        tempappr.append(ma)

        fontRed = Font(color='FF0000', size=16)  # point at red font
        fontBoldRed = Font(color='FF0000', bold=True, size=16)
        if ma > maxsingle:

            # sheet.cell(1+id, cell).font = fontBoldRed  # C4.5
            sheet.cell(3+id, cell).font = fontBoldRed  # SVM

            print('meanDe233', ma)
        if index == 2 or index == 9 or index == 12:
            maxcell = max(tempappr)  # 每個方法的最大值
            if maxcell < maxsingle:
                print("new round3", tempappr)
                maxindex = tempappr.index(maxcell)  # 最大值的 index
            # print('maxindex', maxindex)
                r = 3+id  # SVM
                # r = 1 + id  # C4.5
                c = cell - (len(tempappr) - maxindex) + 1
                print('row3', r, '  column4', c)
                sheet.cell(r, c).font = fontRed

                print('meanDe111', maxcell)
            tempappr = []
        # 若大於 singel method 則標粗體

        wb.save(r'C:\Users\lab723\Desktop\Research\test2-1.xlsx')
    os.chdir(originpath)


'''baseline 跟 single method'''


def singleMethod(train, test, row, column, element, path):
    # element 是指方法
    # 純 polynom_fit_SMOTE
    # 要求出最大值，其他方法要與 single method 的最大值比，如果比 single method 大
    # 則使用 border 做記號
    # 所以這邊要做 return 每個檔案中的 single method 最大值
    tempfileName = train[0]  # 要擷取 file 的名字
    tempfileName = tempfileName.split('-')
    fileName = '-'.join(tempfileName[:-2])
    originpath = os.getcwd()
    os.chdir(path)
    accuraciesDe = []
    accuraciesSVM = []

    '''different dataset '''
    # singlemethod = ['baseline', 'smote', 'poly', 'prow', 'SMOTEIPF']
    start = time.process_time()
    for ii, i in enumerate(train):
        # print("loop", train, element)
        le = preprocessing.LabelEncoder()
        data = pd.read_excel(i, index_col=0)

        classCount, finaldata, output = RandomCal.preprocess(data)
        # 把非 numeric 的資料用 label encoder 轉成 numeric 資料
        for j in range(finaldata.shape[1]):
            for k in range(finaldata.shape[0]):
        # print(df.iloc[j,i])
                if isinstance(finaldata.iloc[k,j],str):
                    finaldata.iloc[:,j] = finaldata.iloc[:,j].apply(lambda col: str(col)) 
                    finaldata.iloc[:,j] = le.fit_transform(finaldata.iloc[:,j])
                    break
        X_polynom, y_polynom = RandomCal.synth(finaldata, output, element)

        # clfDe = DecisionTreeClassifier()
        # clfDe = clfDe.fit(X_polynom, y_polynom)

        clfSVM = svm.SVC(kernel='rbf', C=1, gamma='auto')
        clfSVM = clfSVM.fit(X_polynom, y_polynom)

        # 不然會有多出來的 unnamed column
        test_file = pd.read_excel(test[ii], index_col=0)
        test_data = pd.DataFrame(test_file)

        classCount_test, test_X, test_y = RandomCal.preprocess(test_data)

        for col in range(test_X.shape[1]):
            if isinstance(test_X.iloc[0, :][col], str):
                test_X.iloc[:, col] = le.fit_transform(
                    test_X.iloc[:, col])
        
        le = preprocessing.LabelEncoder()

        # test_y_predicted_De = clfDe.predict(test_X)
        # test_y_predicted_De = le.fit_transform(test_y_predicted_De)

        test_y_predicted_SVM = clfSVM.predict(test_X)
        test_y_predicted_SVM = le.fit_transform(test_y_predicted_SVM)

        # accuracyDe = roc_auc_score(test_y, test_y_predicted_De)
        # accuraciesDe.append(accuracyDe)

        accuracySVM = roc_auc_score(test_y, test_y_predicted_SVM)
        accuraciesSVM.append(accuracySVM)

    # meanDe = statistics.mean(accuraciesDe)
    # meanDe = round(meanDe, 3)
    meanSVM = statistics.mean(accuraciesSVM)
    meanSVM = round(meanSVM, 3)
    end = time.process_time()

    # write in excel prepare
    # 時間
    dur = end - start
    wb = load_workbook(r'C:\Users\lab723\Desktop\Research\test2-1.xlsx')
    # sheet = wb['Ensemble 三個方法 Random']
    sheet = wb['single']
    sheet.cell(row=1+row, column=1, value=fileName)
    # sheet.cell(row=1+row, column=column, value=meanDe)
    sheet.cell(row=4+row, column=column, value=dur)  # need to be modified as 4+row when running SVM
    sheet.cell(row=3+row, column=column, value=meanSVM)
    wb.save(r'C:\Users\lab723\Desktop\Research\test2-1.xlsx')

    # print('C4.5 ', meanDe, '\n', 'SVM ', meanSVM)
    os.chdir(originpath)
    # return meanDe  # 回傳所有 single method 的值
    return meanSVM


if __name__ == "__main__":

    print('123344')
    # f 為三個 methods
    f1 = list(permutations("442", 3))
    f2 = list(permutations("253", 3))
    f3 = list(permutations("334", 3))

    f = f1+f2+f3
    temp = []
    # get rid of the repeat ratio composition
    for i in f:
        if i not in temp:
            temp.append(i)
    f = temp
    imbalDataset = ["imb_IRhigherThan9p3"]
    # array of all folder need to be read, there are train and test excel in it
    folderpath = data_process.get_excel(imbalDataset)
    print('folderpath',folderpath)
    cell = 2
    id = 0
    print(os.getcwd())

    # m 為 2 個 methods
    m1 = list(permutations("550", 3))
    m2 = list(permutations("370", 3))
    m3 = list(permutations("280", 3))

    m = m1+m2+m3
    tempm = []
    # get rid of the repeat ratio composition
    for i in m:
        if i not in tempm:
            tempm.append(i)
    m = tempm

# 下面兩者在執行時  二選一

# Ensemble method

    # for folder in folderpath:

    #     print("folder", folder)

    #     train, test = train_test_split(folder)
    #     print(train)
    #     path = folder
    #     '''
    #     calculatethreemethod(train, test, id, path,
    #                          'elbowCenter')  # Ensemble 3 methods

    #     '''

    #     calculatetwomethod(train, test, id, path, 'random')
    #     id = id+5


# 要執行 single method

    # row = 0
    # for folder in folderpath:
    #     # print("cwd", os.getcwd())
    #     print("folder", folder)

    #     train, test = train_test_split(folder)
    #     print(train)
    #     path = folder
    #     # calculatethreemethod(train, test, id, path,
    #     #                      'elbowRandom')  # Ensemble method
    #     singlemethod = ['baseline', 'smote', 'poly', 'prow', 'SMOTEIPF']
    #     column = 2
    #     for element in singlemethod:
    #         singleMethod(train, test, row, column, element, path)  # 單一方法
    #         column = column + 1

    #     row = row+5

# merge
    singlelist = []
    for folder in folderpath:

        print("folder123", folder)

        train, test = train_test_split(folder)
        print(train)
        path = folder
        '''
            calculatethreemethod(train, test, id, path,
                                'elbowCenter')  # Ensemble 3 methods

            '''
        singlemethod = ['baseline', 'smote', 'poly', 'prow', 'SMOTEIPF']
        column = 2

        for element in singlemethod:
            accursingle = singleMethod(
                train, test, id, column, element, path)  # 單一方法
            singlelist.append(accursingle)
            column = column + 1
        maxsingle = max(singlelist)
        print('max', maxsingle)

        # 三個方法的 Enemble 全部 四種 methods
        # 給single最大值到 calculatetwomethod的參數中
        print('--------- Single Done -------------')
        calculatethreemethod(train, test, id, path, 'random',
                             maxsingle, 'Ensemble 兩個方法 Random')
        print('--------- Random Done -------------')
        calculatethreemethod(train, test, id, path, 'center',
                             maxsingle, 'Ensemble 兩個方法 Center')
        print('--------- Center Done -------------')
        calculatethreemethod(train, test, id, path, 'elbowRandom',
                             maxsingle, 'Ensemble 兩個方法 ElbowRandom')
        print('--------- Elbow Random Done -------------')
        calculatethreemethod(train, test, id, path, 'elbowCenter',
                             maxsingle, 'Ensemble 兩個方法 ElbowCenter')

        """
        兩個方法的 Enemble 全部 四種 methods
        # 給single最大值到 calculatetwomethod的參數中
        print('--------- Single Done -------------')
        calculatetwomethod(train, test, id, path, 'random',
                           maxsingle, 'Ensemble 兩個方法 Random')
        print('--------- Random Done -------------')
        calculatetwomethod(train, test, id, path, 'center',
                           maxsingle, 'Ensemble 兩個方法 Center')
        print('--------- Center Done -------------')
        calculatetwomethod(train, test, id, path, 'elbowRandom',
                           maxsingle, 'Ensemble 兩個方法 ElbowRandom')
        print('--------- Elbow Random Done -------------')
        calculatetwomethod(train, test, id, path, 'elbowCenter',
                           maxsingle, 'Ensemble 兩個方法 ElbowCenter')
                           
        """
        id = id+5
